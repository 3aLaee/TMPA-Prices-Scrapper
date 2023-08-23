import pandas as pd
import locale
from openpyxl import Workbook  # Import the Workbook class
from openpyxl.styles import Font, PatternFill
# Import barchart related classes
from openpyxl.chart import BarChart, Reference, Series


# Set the locale to handle comma as decimal and thousands separator
locale.setlocale(locale.LC_ALL, '')

# Read data from the Excel files
df_aller1 = pd.read_excel('aller1_data.xlsx')
df_aller1V = pd.read_excel('aller1V_data.xlsx')
df_aller4V = pd.read_excel('aller4V_data.xlsx')

# Function to convert formatted price strings to numeric values
def convert_price(price_str):
    return locale.atof(price_str)

# Convert the price columns to numeric format
df_aller1['Price 1PAX'] = df_aller1['Price 1PAX'].apply(convert_price)
df_aller1V['Price 1PAX+1VEH'] = df_aller1V['Price 1PAX+1VEH'].apply(convert_price)
df_aller4V['Price 4PAX+1VEH'] = df_aller4V['Price 4PAX+1VEH'].apply(convert_price)

# Merge dataframes on 'Date' and 'Trajet' columns
merged_df = df_aller1.merge(df_aller1V, on=['Date', 'Trajet'], how='outer')
merged_df = merged_df.merge(df_aller4V, on=['Date', 'Trajet'], how='outer')

# Group by 'Date' and 'Trajet', then calculate the average of each price column
grouped_df = merged_df.groupby(['Date', 'Trajet']).agg({
    'Price 1PAX': 'mean',
    'Price 1PAX+1VEH': 'mean',
    'Price 4PAX+1VEH': 'mean'
}).reset_index()

# Rename columns
grouped_df.columns = ['Date', 'Trajet', 'Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']

# Format numeric columns with appropriate currency style
currency_format = locale.localeconv()['currency_symbol'] + '#,##0.00'
for col in ['Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']:
    grouped_df[col] = grouped_df[col].apply(lambda x: locale.currency(x, symbol=False))

# Write the final table to an Excel file
output_path = 'output_table_styled.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'Average Prices'

# Write the column headers
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
for col_num, col_name in enumerate(grouped_df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    # Set column width based on the length of the column header
    column_letter = chr(64 + col_num)
    column_width = max(len(col_name), len("Average Price 4PAX+1VEH")) + 2
    ws.column_dimensions[column_letter].width = column_width

# Write the data
for row_num, row_data in enumerate(grouped_df.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        # Set column width based on the length of the content
        column_letter = chr(64 + col_num)
        content_width = len(str(value)) + 2
        if content_width > ws.column_dimensions[column_letter].width:
            ws.column_dimensions[column_letter].width = content_width

# Format numeric columns with appropriate currency style
currency_format = locale.localeconv()['currency_symbol'] + '#,##0.00'
for col in ['Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']:
    for row in ws.iter_rows(min_col=grouped_df.columns.get_loc(col) + 1, max_col=grouped_df.columns.get_loc(col) + 1,
                            min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.number_format = currency_format


# Create a new worksheet for the grouped bar chart
ws_price_comparison = wb.create_sheet(title='Price Comparison by Route')

# Add data to the chart
chart = BarChart()
chart.title = "Price Comparison by Route"
chart.y_axis.title = "Average Price"
chart.x_axis.title = "Route"

company_names = ['FRS', 'Naviera Armas Trasmediterranea', 'Balearia', 'AML - Africa Moroco Link']

for company_name in company_names:
    for ticket_type in ['Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']:
        data = Reference(ws, min_col=grouped_df.columns.get_loc(ticket_type) + 1,
                         min_row=2, max_row=grouped_df.shape[0] + 1,
                         max_col=grouped_df.columns.get_loc(ticket_type) + 1)
        series = Series(data, title=f'{company_name} - {ticket_type}')
        chart.append(series)

# Set category labels (routes)
categories = Reference(ws, min_col=2, min_row=2, max_row=grouped_df.shape[0] + 1)
chart.set_categories(categories)

# Add the chart to the new worksheet
ws_price_comparison.add_chart(chart, "A1")


# Save the workbook
wb.save(output_path)

# Provide feedback to the user
print(f"Styled and formatted data has been saved to {output_path}")
