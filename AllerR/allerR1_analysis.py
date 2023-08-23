import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import PatternFill
from io import BytesIO
import matplotlib.pyplot as plt

# Read the Excel file
file_path = 'allerR1_data.xlsx'
df = pd.read_excel(file_path)

# Convert the 'Date' column to datetime format
df['Date'] = pd.to_datetime(df['Date']).dt.date  # Extract only the date part

# Convert the 'Price' column to numeric format with French number format
df['Price 1PAX'] = pd.to_numeric(df['Price 1PAX'].str.replace('.', '').str.replace(',', '.'), errors='coerce')

# Group by 'Trajet', 'Company', and 'Date' and calculate summary statistics
summary = df.groupby(['Trajet', 'Company', 'Date'])['Price 1PAX'].agg(['mean', 'min', 'max'])

# Create a new Excel file and add summary statistics
output_file = 'allerR1_analysis.xlsx'
workbook = Workbook()

# Create summary sheet
summary_sheet = workbook.active
summary_sheet.title = 'Summary'
summary_sheet.append(['Trajet', 'Company', 'Date', 'Average Price', 'Min Price', 'Max Price'])

for index, row in summary.iterrows():
    summary_sheet.append([index[0], index[1], index[2], row['mean'], row['min'], row['max']])

# Analyze prices and create charts for each 'Trajet'
for trajet in df['Trajet'].unique():
    trajet_df = df[df['Trajet'] == trajet]

    # Create a new sheet for the 'Trajet' analysis
    trajet_sheet = workbook.create_sheet(title=trajet)

    # Create a line chart for price progression by company
    for company in trajet_df['Company'].unique():
        company_df = trajet_df[trajet_df['Company'] == company]
        line_chart = LineChart()
        line_chart.title = f'Price Progression - {trajet} - {company}'
        line_chart.x_axis.title = 'Date'
        line_chart.y_axis.title = 'Price 1PAX'

        dates = Reference(summary_sheet, min_col=3, min_row=2, max_row=summary_sheet.max_row)
        prices = Reference(trajet_sheet, min_col=3, max_col=3, min_row=1, max_row=trajet_df.shape[0] + 1)
        line_chart.add_data(prices, titles_from_data=True)
        line_chart.set_categories(dates)

        trajet_sheet.add_chart(line_chart, f"A{(len(trajet_df['Company'].unique()) + 2) * (trajet_df['Company'].unique().tolist().index(company) + 1)}")

# Create a bar chart for average prices by company and 'Trajet'
bar_chart_sheet = workbook.create_sheet(title='Average Prices')
bar_chart = BarChart()
bar_chart.title = 'Average Prices by Company and Trajet'
bar_chart.x_axis.title = 'Company'
bar_chart.y_axis.title = 'Average Price'
bar_chart.y_axis.majorGridlines = None

avg_prices = Reference(summary_sheet, min_col=4, max_col=4, min_row=2, max_row=summary_sheet.max_row)
companies = Reference(summary_sheet, min_col=2, min_row=2, max_row=summary_sheet.max_row)
bar_chart.add_data(avg_prices, titles_from_data=True)
bar_chart.set_categories(companies)

bar_chart_sheet.add_chart(bar_chart, 'A2')

# Save the workbook
workbook.save(output_file)
print(f"Excel file '{output_file}' created.")
