import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart3D, Reference
import locale
from openpyxl.chart import (
    BarChart,
    LineChart,
    AreaChart,
    Series,
    PieChart,

)


# Read data from the Excel files
df_aller1 = pd.read_excel('aller1_data.xlsx')
df_aller1V = pd.read_excel('aller1V_data.xlsx')
df_aller4V = pd.read_excel('aller4V_data.xlsx')

# Function to convert formatted price strings to numeric values
def convert_price(price_str):
    return float(price_str.replace(',', '.'))  # Replace comma with dot for decimal

# Convert the price columns to numeric format
df_aller1['Price 1PAX'] = df_aller1['Price 1PAX'].apply(convert_price)
df_aller1V['Price 1PAX+1VEH'] = df_aller1V['Price 1PAX+1VEH'].apply(convert_price)
df_aller4V['Price 4PAX+1VEH'] = df_aller4V['Price 4PAX+1VEH'].apply(convert_price)

# Merge dataframes on 'Date' and 'Trajet' columns
merged_df = df_aller1.merge(df_aller1V, on=['Date', 'Trajet'], how='outer')
merged_df = merged_df.merge(df_aller4V, on=['Date', 'Trajet'], how='outer')

# Group by 'Date' and 'Trajet', then calculate the average of each price column
grouped_df = merged_df.groupby(['Date', 'Trajet']).agg({
    'Price 1PAX': 'mean',
    'Price 1PAX+1VEH': 'mean',
    'Price 4PAX+1VEH': 'mean'
}).reset_index()

# Rename columns
grouped_df.columns = ['Date', 'Trajet', 'Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = 'Average Prices'

# Write the column headers for the first sheet
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
for col_num, col_name in enumerate(grouped_df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    # Set column width based on the length of the column header
    column_letter = chr(64 + col_num)
    column_width = max(len(col_name), len("Average Price 4PAX+1VEH")) + 2
    ws.column_dimensions[column_letter].width = column_width

# Write the data for the first sheet
for row_num, row_data in enumerate(grouped_df.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        # Set column width based on the length of the content
        column_letter = chr(64 + col_num)
        content_width = len(str(value)) + 2
        if content_width > ws.column_dimensions[column_letter].width:
            ws.column_dimensions[column_letter].width = content_width

# Format numeric columns with appropriate currency style for the first sheet
for col in ['Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']:
    for row in ws.iter_rows(min_col=grouped_df.columns.get_loc(col) + 1, max_col=grouped_df.columns.get_loc(col) + 1,
                            min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.number_format = locale.localeconv()['currency_symbol'] + '#,##0.00'

# Calculate cheapest and most expensive companies for each price column
cheapest_price_1pax = merged_df.groupby(['Date', 'Trajet'])['Price 1PAX'].idxmin()
expensive_price_1pax = merged_df.groupby(['Date', 'Trajet'])['Price 1PAX'].idxmax()
cheapest_price_1pax_1veh = merged_df.groupby(['Date', 'Trajet'])['Price 1PAX+1VEH'].idxmin()
expensive_price_1pax_1veh = merged_df.groupby(['Date', 'Trajet'])['Price 1PAX+1VEH'].idxmax()
cheapest_price_4pax_1veh = merged_df.groupby(['Date', 'Trajet'])['Price 4PAX+1VEH'].idxmin()
expensive_price_4pax_1veh = merged_df.groupby(['Date', 'Trajet'])['Price 4PAX+1VEH'].idxmax()

# Create a DataFrame with the cheapest and expensive companies information
cheapest_expensive_df = pd.DataFrame({
    'Date': merged_df.loc[cheapest_price_1pax, 'Date'],
    'Trajet': merged_df.loc[cheapest_price_1pax, 'Trajet'],
    'Cheapest Company 1PAX': merged_df.loc[cheapest_price_1pax, 'Company_x'],
    'Expensive Company 1PAX': merged_df.loc[expensive_price_1pax, 'Company_x'],
    'Cheapest Company 1PAX+1VEH': merged_df.loc[cheapest_price_1pax_1veh.dropna(), 'Company_y'],
    'Expensive Company 1PAX+1VEH': merged_df.loc[expensive_price_1pax_1veh.dropna(), 'Company_y'],
    'Cheapest Company 4PAX+1VEH': merged_df.loc[cheapest_price_4pax_1veh, 'Company'],
    'Expensive Company 4PAX+1VEH': merged_df.loc[expensive_price_4pax_1veh, 'Company']
})

# Create a new worksheet for the cheapest and expensive companies data
ws_cheapest_expensive = wb.create_sheet(title='Cheapest & Expensive Companies')

# Write the column headers for the second sheet
for col_num, col_name in enumerate(cheapest_expensive_df.columns, 1):
    cell = ws_cheapest_expensive.cell(row=1, column=col_num, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Write the data for the second sheet
for row_num, row_data in enumerate(cheapest_expensive_df.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_cheapest_expensive.cell(row=row_num, column=col_num, value=value)

# Get unique trajet values
unique_trajets = grouped_df['Trajet'].unique()

# Create a new worksheet for the charts
ws_charts = wb.create_sheet(title='Trajet Charts')

# Loop through each unique trajet and create a chart
for index, trajet in enumerate(unique_trajets, start=1):
    # Filter the data for the current trajet
    trajet_data = grouped_df[grouped_df['Trajet'] == trajet]

    # Create a new 3D Column Chart
    chart = BarChart3D()
    chart.title = f"Price Comparison for Trajet: {trajet}"
    chart.y_axis.title = "Price"
    chart.x_axis.title = "Date"

    # Set the minimum and maximum values for the y-axis (price)
    chart.y_axis.scaling.min = 10  # Minimum value for the y-axis
    chart.y_axis.scaling.max = None  # The maximum value will be automatically determined

    # Add data to the chart
    data = Reference(ws_charts, min_col=3, min_row=index * (len(trajet_data) + 1) + 1,
                     max_col=5, max_row=(index + 1) * (len(trajet_data) + 1) - 1)
    categories = Reference(ws_charts, min_col=2, min_row=index * (len(trajet_data) + 1) + 1,
                           max_row=(index + 1) * (len(trajet_data) + 1) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Add the chart to the charts worksheet
    ws_charts.add_chart(chart, f"A{index * (len(trajet_data) + 1)}")
    
# ... (existing code)

# ... (existing code)

# Create a new worksheet for the grouped bar chart
ws_price_comparison = wb.create_sheet(title='Price Comparison by Route')

# Add data to the chart
chart = BarChart()
chart.title = "Price Comparison by Route"
chart.y_axis.title = "Average Price"
chart.x_axis.title = "Route"

company_names = ['FRS', 'Naviera Armas Trasmediterranea', 'Balearia', 'AML - Africa Moroco Link']

for company_name in company_names:
    for ticket_type in ['Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']:
        data = Reference(ws, min_col=grouped_df.columns.get_loc(ticket_type) + 1,
                         min_row=2, max_row=grouped_df.shape[0] + 1,
                         max_col=grouped_df.columns.get_loc(ticket_type) + 1)
        series = Series(data, title=f'{company_name} - {ticket_type}')
        chart.append(series)

# Set category labels (routes)
categories = Reference(ws, min_col=2, min_row=2, max_row=grouped_df.shape[0] + 1)
chart.set_categories(categories)

# Add the chart to the new worksheet
ws_price_comparison.add_chart(chart, "A1")

# ... (continue with other charts)

# ... (existing code)

# ... (existing code)

# Create a new worksheet for the stacked area chart
ws_company_performance = wb.create_sheet(title='Company Performance')

# Add data to the chart
chart = AreaChart()
chart.title = "Company Performance"
chart.y_axis.title = "Price"
chart.x_axis.title = "Date"

company_names = ['FRS', 'Naviera Armas Trasmediterranea', 'Balearia', 'AML - Africa Moroco Link']

for company_name in company_names:
    data = Reference(ws, min_col=grouped_df.columns.get_loc('Average Price 1PAX') + 3,
                     min_row=2, max_row=grouped_df.shape[0] + 1,
                     max_col=grouped_df.columns.get_loc('Average Price 1PAX') + 3)
    series = Series(data, title=company_name)
    chart.append(series)

# Set category labels (dates)
categories = Reference(ws, min_col=1, min_row=2, max_row=grouped_df.shape[0] + 1)
chart.set_categories(categories)

# Set chart style to stacked area
chart.grouping = "stacked"

# Add the chart to the new worksheet
ws_company_performance.add_chart(chart, "A1")

# ... (continue with other charts)


# Create a new worksheet for the stacked area chart
ws_company_performance = wb.create_sheet(title='Company Performance')

# Add data to the chart
chart = AreaChart()
chart.title = "Company Performance"
chart.y_axis.title = "Price"
chart.x_axis.title = "Date"

for col_num, col_name in enumerate(grouped_df.columns[2:], 1):
    data = Reference(ws, min_col=col_num + 2, min_row=2, max_row=grouped_df.shape[0] + 1)
    series = Series(data, title_from_data=True)
    chart.append(series)

# Set category labels (dates)
categories = Reference(ws, min_col=1, min_row=2, max_row=grouped_df.shape[0] + 1)
chart.set_categories(categories)

# Set chart style to stacked area
chart.grouping = "stacked"

# Add the chart to the new worksheet
ws_company_performance.add_chart(chart, "A1")

# ... (continue with other charts)

# ... (existing code)
# Create a new worksheet for the line chart
ws_average_prices = wb.create_sheet(title='Average Prices Over Time')

# Create a LineChart
line_chart = LineChart()
line_chart.title = "Average Prices Over Time"
line_chart.y_axis.title = "Average Price"
line_chart.x_axis.title = "Date"

# Add data to the chart
for company_type in ['Average Price 1PAX', 'Average Price 1PAX+1VEH', 'Average Price 4PAX+1VEH']:
    data = Reference(ws, min_col=grouped_df.columns.get_loc(company_type) + 3,
                     min_row=2, max_row=grouped_df.shape[0] + 1,
                     max_col=grouped_df.columns.get_loc(company_type) + 3)
    series = Series(data, title=company_type)
    line_chart.append(series)

# Set category labels (dates)
categories = Reference(ws, min_col=1, min_row=2, max_row=grouped_df.shape[0] + 1)
line_chart.set_categories(categories)

# Add the chart to the new worksheet
ws_average_prices.add_chart(line_chart, "A1")


    

# Save the workbook
output_path = 'output_charts.xlsx'
wb.save(output_path)

# Provide feedback to the user
print(f"Charts have been saved to {output_path}")
